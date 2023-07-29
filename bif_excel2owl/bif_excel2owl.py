# -*- coding: utf-8 -*-
"""BIF_Exel2OWL.py
    USE: python bif_exel2owl.py config bifd_url infile outfile
"""
import openpyxl
from owlready2 import *
import inspect
import json


def createCircuits(ws, config):
    for p in range(2):
        for i in range(ws.max_row - 1):
            val = ws.cell(row=i + 2, column=config["ID"]).value
            if val is not None and val.strip() != "":
                className = val.strip().replace(' ', '_')
                if p == 0:  # 1st pass
                    if val.find(':') < 0:
                        functionality = ws.cell(row=i + 2, column=config["Functionality"]).value  # Functionality
                        references = ws.cell(row=i + 2, column=config["References"]).value  # References
                        uv = ws.cell(row=i + 2, column=config["Uniform"]).value  # Uniform?
                        uniform = False
                        if uv is not None and uv:
                            uniform = True
                        with onto:
                            try:
                                eval('onto.' + className)
                            except SyntaxError:
                                className = '_' + className
                            if uniform:
                                cls = types.new_class(className, (bifdns.UniformCircuit,))
                            else:
                                cls = types.new_class(className, (bifdns.Circuit,))
                            labels = ws.cell(row=i + 2, column=config["Names"]).value
                            addLabels(labels, cls)
                            if functionality is not None and functionality != "":
                                cls.functionality = functionality
                            if references is not None and references != "":
                                references = references.split(";")
                                cls.reference = references
                else:   # 2nd pass
                    parts = ws.cell(row=i + 2, column=config["Parts"]).value
                    superClasses = ws.cell(row=i + 2, column=config["SuperClasses"]).value
                    addHasPart(parts, className)
                    addSubClassOf(superClasses, className)


def addLabels(labels, cls):
    if labels:
        lbls = labels.split(";")
        for label in lbls:
            cls.label.append(label)


def addHasPart(parts, className):
    if parts:
        prts = parts.split(";")
    else:
        return
    pos = className.find(':')
    if pos < 0:
        try:
            cls = eval("onto." + className)
        except SyntaxError:
            className = "_" + className
            cls = eval("onto." + className)
        for part in prts:
            pn = part.strip()
            pos = pn.find(':')
            if pos < 0:
                pn = pn.replace(' ', '_')
                try:
                    prt = eval("onto." + pn)
                except SyntaxError:
                    pn = '_' + pn
                    prt = eval("onto." + pn)
            else:
                prefix = pn[:pos]
                pn = pn[pos + 1:].replace(' ', '_')
                ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
                with ex_onto:
                    prt = types.new_class(pn, (bifdns.Circuit,))
            if prt is not None and inspect.isclass(prt):
                cls.is_a.append(bifdns.hasPart.some(prt))
    else:
        prefix = className[:pos]
        className = className[pos + 1:].replace(' ', '_')
        ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
        with ex_onto:
            cls = types.new_class(className, (bifdns.Circuit,))
        for part in prts:
            pn = part.strip()
            pos = pn.find(':')
            if pos < 0:  # no relation is defined for exterior circuits.
                pn = pn.replace(' ', '_')
                prt = eval("onto." + pn)
            if prt is not None and inspect.isclass(prt):
                prt.is_a.append(bifdns.partOf.some(cls))


def addSubClassOf(superClasses, className):
    if superClasses:
        sups = superClasses.split(";")
    else:
        return
    pos = className.find(':')
    if pos < 0:
        cls = eval("onto." + className)
        for superClassName in sups:
            scn = superClassName.strip()
            pos = scn.find(':')
            if pos < 0:
                scn = scn.replace(' ', '_')
                sc = eval("onto." + scn.strip())
            else:
                prefix = scn[:pos]
                ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
                with ex_onto:
                    sc = types.new_class(scn[pos+1:].strip().replace(' ', '_'), (bifdns.Circuit,))
            if sc is not None and inspect.isclass(sc):
                cls.is_a.append(sc)


def createConnections(ws, config):
    for i in range(ws.max_row - 1):
        inputCircuit = ""
        input = ws.cell(row=i + 2, column=config["InputCircuit"]).value
        if input is not None and input.strip() != "":
            pos = input.find(':')
            inputCircuit = input.strip().replace(' ', '_') if pos < 0 else input[pos + 1:].strip().replace(' ', '_')
        outputCircuit = ""
        output = ws.cell(row=i + 2, column=config["OutputCircuit"]).value
        if output is not None and output.strip() != "":
            pos = output.find(':')
            outputCircuit = output.strip().replace(' ', '_') if pos < 0 else output[pos + 1:].strip().replace(' ', '_')
        if inputCircuit == "" or outputCircuit == "":
            continue
        connectionID = inputCircuit + "-" + outputCircuit
        with onto:
            cls = types.new_class(connectionID, (bifdns.Connection,))
            cls.label = connectionID
            addInOut(cls, input.strip(), output.strip())
            functionality = ws.cell(row=i + 2, column=config["Functionality"]).value  # Functionality
            references = ws.cell(row=i + 2, column=config["References"]).value  # References
            if functionality is not None and functionality != "":
                cls.functionality = functionality
            if references is not None and references != "":
                references = references.split(";")
                cls.reference = references


def addInOut(cls, inputCircuit, outputCircuit):
    pos = inputCircuit.find(':')
    if pos < 0:
        inputClass = eval("onto." + inputCircuit.replace(' ', '_'))
    else:
        prefix = inputCircuit[:pos].strip().replace(' ', '_')
        inputCircuit = inputCircuit[pos + 1:].strip().replace(' ', '_')
        ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
        with ex_onto:
            inputClass = types.new_class(inputCircuit, (bifdns.Circuit,))
    cls.is_a.append(bifdns.inputCircuit.some(inputClass))
    pos = outputCircuit.find(':')
    if pos < 0:
        outputClass = eval("onto." + outputCircuit.replace(' ', '_'))
    else:
        prefix = outputCircuit[:pos].strip().replace(' ', '_')
        outputCircuit = outputCircuit[pos + 1:].strip().replace(' ', '_')
        ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
        with ex_onto:
            outputClass = types.new_class(outputCircuit, (bifdns.Circuit,))
    cls.is_a.append(bifdns.outputCircuit.some(outputClass))


if __name__ == '__main__':
    if len(sys.argv) <= 3:
        print("USE: python bif_exel2owl.py config bifd_url infile outfile")
        exit()

    configPath = sys.argv[1]
    infilePath = sys.argv[3]
    outfilePath = sys.argv[4]
    lastSlash = infilePath.rfind('/')
    if lastSlash >= 0:
        infileName = infilePath[lastSlash + 1:]
    else:
        infileName = infilePath

    with open(configPath) as f:
        config = json.load(f)

    wb = openpyxl.load_workbook(infilePath)
    bifd = get_ontology(sys.argv[2]).load()

    bifdns = bifd.get_namespace('https://wba-initiative.org/bifd/')
    # Defining annotation properties
    with bifdns:
        class functionality(comment):
            pass


        class reference(comment):
            pass


        class implementation(comment):
            pass

    # Defining an ontoloby
    project = wb['Project']
    project_name_column = config["Project"]["DataSet"]
    pname = project.cell(row=2, column=project_name_column).value
    if not pname:
        print("Error: no project name")
        exit()
    description = project.cell(row=2, column=config["Project"]["Description"]).value
    onto = get_ontology("https://wba-initiative.org/wbra/" + pname + "/")
    onto.metadata.comment.append(description)

    createCircuits(wb['Circuit'], config["Circuit"])
    createConnections(wb['Connection'], config["Connection"])
    # onto_path.append(outfilePath)
    onto.save(file=outfilePath)

# -*- coding: utf-8 -*-
"""BIF_Exel2OWL.py
    USE: python bif_exel2owl.py config bifd_url infile outfile
"""
import openpyxl
from rdflib import Graph, URIRef, Literal, BNode, Namespace
from rdflib.namespace import RDFS, RDF, OWL
import json
import sys

DOIs = {}
URLs = {}


def references(ws, config):
    for i in range(ws.max_row - 1):
        id = ws.cell(row=i + 2, column=config["ID"]).value
        if id is not None and id.strip()!="":
            id = id.strip()
            doi = ws.cell(row=i + 2, column=config["DOI"]).value
            if doi is not None and doi.strip()!="":
                DOIs[id] = doi.strip()
            url = ws.cell(row=i + 2, column=config["URL"]).value
            if url is not None and url.strip() != "":
                URLs[id] = url.strip()


def createCircuits(ws, config):
    for i in range(ws.max_row - 1):
        projectID = None
        if "ProjectID" in config:
            projectID = ws.cell(row=i + 2, column=config["ProjectID"]).value
            if projectID is not None and projectID.strip() != "":
                projectID = projectID.strip()
        val = ws.cell(row=i + 2, column=config["ID"]).value
        if val is not None and val.strip() != "":
            id = val.strip().replace(' ', '_')
            pos = id.find(':')
            if pos < 0:
                circuit = URIRef(onto_URI + id)
            else:
                prefix = id[:pos]
                className = id[pos+1:]
                uri = "https://wba-initiative.org/wbra/" + prefix + "/"
                g.bind(prefix, Namespace(URIRef(uri)))
                circuit = URIRef(uri + className)
            g.add((circuit, RDF.type, OWL.Class))
            if "Source" in config:
                source = ws.cell(row=i + 2, column=config["Source"]).value
                if source is not None and source.strip() != "":
                    source = source.strip()
                    if source[0] == '[':
                        rp = referencePointer(source)
                        if rp is not None:
                            g.add((circuit, BIFD.reference, Literal(rp)))
                        else:
                            print("Reference pointer not found:" + source, file=sys.stderr)
                        source = projectID
                    elif source == "makeshift" or source == "collection":
                        source = projectID
                    if source is not None:
                        g.add((circuit, BIFD.source, Literal(source)))
            references = None
            if "References" in config:
                references = ws.cell(row=i + 2, column=config["References"]).value  # References
            uv = ws.cell(row=i + 2, column=config["Uniform"]).value  # Uniform?
            uniform = False
            if uv is not None and uv:
                uniform = True
            if uniform:
                g.add((circuit, RDFS.subClassOf, BIFD.UniformCircuit))
            else:
                g.add((circuit, RDFS.subClassOf, BIFD.Circuit))
            labels = ws.cell(row=i + 2, column=config["Names"]).value
            if labels is None:
                labels = id
            labels = labels.split(";")
            for label in labels:
                g.add((circuit, RDFS.label, Literal(label)))
            if "Functionality" in config:
                functionality = ws.cell(row=i + 2, column=config["Functionality"]).value  # Functionality
                if functionality is not None and functionality != "":
                    g.add((circuit, BIFD.functionality, Literal(functionality)))
            if "OutputSemantics" in config:
                outputSemantics = ws.cell(row=i + 2, column=config["OutputSemantics"]).value
                if outputSemantics is not None and outputSemantics != "":
                    g.add((circuit, BIFD.outputSemantics, Literal(outputSemantics)))
            if references is not None and references != "":
                references = references.split(";")
                for reference in references:
                    rp = referencePointer(reference)
                    if rp is not None:
                        g.add((circuit, BIFD.reference, Literal(rp)))
                    else:
                        print("Reference pointer not found:" + reference, file=sys.stderr)
            parts = ws.cell(row=i + 2, column=config["Parts"]).value
            if parts:
                addHasPart(parts, circuit)
            superClasses = ws.cell(row=i + 2, column=config["SuperClasses"]).value
            if superClasses:
                addSubClassOf(superClasses, circuit)


def addSome(x, y, property):
    restriction = BNode()
    g.add((restriction, RDF.type, OWL.Restriction))
    g.add((restriction, OWL.onProperty, property))
    g.add((restriction, OWL.someValuesFrom, y))
    g.add((x, RDFS.subClassOf, restriction))


def addHasPart(parts, circuit):
    prts = parts.split(";")
    for part in prts:
        pn = part.strip().replace(' ', '_')
        pos = pn.find(':')
        if pos < 0:
            partID = pn
            part = URIRef(onto_URI + partID)
        else:
            prefix = pn[:pos]
            partID = pn[pos+1:]
            uri = "https://wba-initiative.org/wbra/" + prefix + "/" + partID
            part = URIRef(uri)
        addSome(circuit, part, BIFD.hasPart)


def addSubClassOf(superClasses, circuit):
    sups = superClasses.split(";")
    for superClassName in sups:
        scn = superClassName.strip().replace(' ', '_')
        pos = scn.find(':')
        if pos < 0:
            scID = scn
            sc = URIRef(onto_URI + scID)
        else:
            prefix = scn[:pos]
            scID = scn[pos+1:]
            uri = "https://wba-initiative.org/wbra/" + prefix + "/" + scID
            sc = URIRef(uri)
        g.add((circuit, RDFS.subClassOf, sc))


def createConnections(ws, config):
    for i in range(ws.max_row - 1):
        inputCircuit = ""
        input = ws.cell(row=i + 2, column=config["InputCircuit"]).value
        if input is not None and input.strip() != "":
            pos = input.find(':')
            inputCircuit = input.strip().replace(' ', '_') if pos < 0 else input[pos + 1:].strip().replace(' ', '_')
        outputCircuit = ""
        output = ws.cell(row=i + 2, column=config["OutputCircuit"]).value
        if output is not None and output.strip() != "":
            pos = output.find(':')
            outputCircuit = output.strip().replace(' ', '_') if pos < 0 else output[pos + 1:].strip().replace(' ', '_')
        if inputCircuit == "" or outputCircuit == "":
            continue
        connectionID = inputCircuit + "-" + outputCircuit
        connection = URIRef(onto_URI + connectionID)
        g.add((connection, RDF.type, OWL.Class))
        g.add((connection, RDFS.subClassOf, BIFD.Connection))
        g.add((connection, RDFS.label, Literal(connectionID)))
        addInOut(connection, input, output)
        functionality = None
        if "Functionality" in config:
            functionality = ws.cell(row=i + 2, column=config["Functionality"]).value  # Functionality
        references = ws.cell(row=i + 2, column=config["References"]).value  # References
        if functionality is not None and functionality != "":
            g.add((connection, BIFD.functionality, Literal(functionality)))
        if references is not None and references != "":
            references = references.split(";")
            for reference in references:
                rp = referencePointer(reference)
                if rp is not None:
                    g.add((connection, BIFD.reference, Literal(rp)))
                else:
                    print("Reference pointer not found:" + reference, file=sys.stderr)


def addInOut(connection, inputCircuit, outputCircuit):
    inputCircuit = inputCircuit.strip().replace(' ', '_')
    pos = inputCircuit.find(':')
    if pos < 0:
        ic = URIRef(onto_URI + inputCircuit)
    else:
        prefix = inputCircuit[:pos].strip()
        icID = inputCircuit[pos + 1:].strip()
        ic = URIRef("https://wba-initiative.org/wbra/" + prefix + "/" + icID)
    addSome(connection, ic, BIFD.inputCircuit)
    outputCircuit = outputCircuit.strip().replace(' ', '_')
    pos = outputCircuit.find(':')
    if pos < 0:
        oc = URIRef(onto_URI + outputCircuit)
    else:
        prefix = outputCircuit[:pos].strip()
        ocID = outputCircuit[pos + 1:].strip()
        oc = URIRef("https://wba-initiative.org/wbra/" + prefix + "/" + ocID)
    addSome(connection, oc, BIFD.outputCircuit)


def referencePointer(ref):
    ref = ref.strip().strip(']').lstrip('[')
    if ref in DOIs:
        return DOIs[ref]
    elif ref in URLs:
        return URLs[ref]
    else:
        return None


if __name__ == '__main__':
    if len(sys.argv) <= 3:
        print("USE: python bif_exel2owl.py config bifd_url infile outfile")
        exit()

    configPath = sys.argv[1]
    bifdPath = sys.argv[2]
    infilePath = sys.argv[3]
    outfilePath = sys.argv[4]
    lastSlash = infilePath.rfind('/')
    if lastSlash >= 0:
        infileName = infilePath[lastSlash + 1:]
    else:
        infileName = infilePath

    with open(configPath) as f:
        config = json.load(f)

    wb = openpyxl.load_workbook(infilePath, data_only=True)
    g = Graph()
    bifd = Graph()
    bifd.parse(bifdPath)
    # getting the base url
    for prefix, namespace in bifd.namespaces():
        if prefix == "":
            base = namespace
    BIFD = Namespace(base)
    g.bind("bifd", BIFD)

    # Defining the ontology
    if 'Project' in wb:
        project = wb['Project']
        project_name_column = config["Project"]["DataSet"]
        pname = project.cell(row=2, column=project_name_column).value
        if not pname:
            print("Error: no project name")
            exit()
        description = project.cell(row=2, column=config["Project"]["Description"]).value
        onto_URI = "https://wba-initiative.org/wbra/" + pname + "/"
        onto = URIRef(onto_URI)
        g.bind(pname, Namespace(onto_URI))
        g.add((onto, RDF.type, OWL.Ontology))
        references(wb['Reference'], config["Reference"])
        createCircuits(wb['Circuit'], config["Circuit"])
        createConnections(wb['Connection'], config["Connection"])
    else:   # WholeBIF
        description = "Whole Brain Information Flow Diagram"    # TBD
        onto_URI = "https://wba-initiative.org/wbra/wholeBIF/"
        g.bind("WholeBIF", Namespace(onto_URI))
        onto = URIRef(onto_URI)
        references(wb['wbReferences'], config["Reference"])
        createCircuits(wb['wbCircuits'], config["Circuit"])
        createConnections(wb['wbConnections'], config["Connection"])
    g.add((onto, RDFS.comment, Literal(description)))
    v = g.serialize(format="pretty-xml")
    f = open(outfilePath, 'w')
    f.write(v)

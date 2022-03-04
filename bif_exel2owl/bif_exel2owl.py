# -*- coding: utf-8 -*-
"""BIF_Exel2OWL.py
    USE: python bif_exel2owl.py infile bifd_url outfile
"""
import openpyxl
from owlready2 import *
import inspect


def createCircuits(ws):
    for p in range(2):
        for i in range(ws.max_row - 1):
            val = ws.cell(row=i + 2, column=1).value
            if val is not None and val.strip() != "":
                className = val.strip().replace(' ', '_')
                if p == 0:  # 1st pass
                    if val.find(':') < 0:
                        clm4 = ws.cell(row=i + 2, column=5).value  # Functionality
                        clm5 = ws.cell(row=i + 2, column=6).value  # References
                        clm6 = ws.cell(row=i + 2, column=7).value  # TODO: implementation
                        clm7 = ws.cell(row=i + 2, column=8).value  # Uniform?
                        uniform = False
                        if clm7 is not None and clm7:
                            uniform = True
                        with onto:
                            if uniform:
                                cls = types.new_class(className, (bifdns.UniformCircuit,))
                            else:
                                cls = types.new_class(className, (bifdns.Circuit,))
                            labels = ws.cell(row=i + 2, column=2).value
                            addLabels(labels, cls)
                            if clm4 is not None and clm4 != "":
                                cls.functionality = clm4
                            if clm5 is not None and clm5 != "":
                                references = clm5.split(";")
                                cls.reference = references
                else:   # 2nd pass
                    parts = ws.cell(row=i + 2, column=4).value
                    superClasses = ws.cell(row=i + 2, column=3).value
                    addHasPart(parts, className)
                    addSubClassOf(superClasses, className)


def addLabels(labels, cls):
    if labels:
        lbls = labels.split(";")
        for label in lbls:
            cls.label.append(label)


def addHasPart(parts, className):
    if parts:
        prts = parts.split(";")
    else:
        return
    pos = className.find(':')
    if pos < 0:
        cls = eval("onto." + className)
        for part in prts:
            pn = part.strip()
            pos = pn.find(':')
            if pos < 0:
                pn = pn.replace(' ', '_')
                prt = eval("onto." + pn)
            else:
                prefix = pn[:pos]
                pn = pn[pos + 1:].replace(' ', '_')
                ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
                with ex_onto:
                    prt = types.new_class(pn, (bifdns.Circuit,))
            if prt is not None and inspect.isclass(prt):
                cls.is_a.append(bifdns.hasPart.some(prt))
    else:
        prefix = className[:pos]
        className = className[pos + 1:].replace(' ', '_')
        ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
        with ex_onto:
            cls = types.new_class(className, (bifdns.Circuit,))
        for part in prts:
            pn = part.strip()
            pos = pn.find(':')
            if pos < 0:  # no relation is defined for exterior circuits.
                pn = pn.replace(' ', '_')
                prt = eval("onto." + pn)
            if prt is not None and inspect.isclass(prt):
                prt.is_a.append(bifdns.partOf.some(cls))


def addSubClassOf(superClasses, className):
    if superClasses:
        sups = superClasses.split(";")
    else:
        return
    pos = className.find(':')
    if pos < 0:
        cls = eval("onto." + className)
        for superClassName in sups:
            scn = superClassName.strip()
            pos = scn.find(':')
            if pos < 0:
                scn = scn.replace(' ', '_')
                sc = eval("onto." + scn.strip())
            else:
                prefix = scn[:pos]
                ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
                with ex_onto:
                    sc = types.new_class(scn[pos+1:].strip().replace(' ', '_'), (bifdns.Circuit,))
            if sc is not None and inspect.isclass(sc):
                cls.is_a.append(sc)


def createConnections(ws):
    for i in range(ws.max_row - 1):
        inputCircuit = ""
        col1 = ws.cell(row=i + 2, column=1).value
        if col1 is not None and col1.strip() != "":
            pos = col1.find(':')
            inputCircuit = col1.strip().replace(' ', '_') if pos < 0 else col1[pos + 1:].strip().replace(' ', '_')
        outputCircuit = ""
        col2 = ws.cell(row=i + 2, column=2).value
        if col2 is not None and col2.strip() != "":
            pos = col2.find(':')
            outputCircuit = col2.strip().replace(' ', '_') if pos < 0 else col2[pos + 1:].strip().replace(' ', '_')
        if inputCircuit == "" or outputCircuit == "":
            continue
        connectionID = inputCircuit + "-" + outputCircuit
        with onto:
            cls = types.new_class(connectionID, (bifdns.Connection,))
            cls.label = connectionID
            addInOut(cls, col1.strip(), col2.strip())
            clm3 = ws.cell(row=i + 2, column=4).value  # TODO: Transmitter
            clm4 = ws.cell(row=i + 2, column=4).value  # Functionality
            clm5 = ws.cell(row=i + 2, column=5).value  # References
            if clm4 is not None and clm4 != "":
                cls.functionality = clm4
            if clm5 is not None and clm5 != "":
                references = clm5.split(";")
                cls.reference = references


def addInOut(cls, inputCircuit, outputCircuit):
    pos = inputCircuit.find(':')
    if pos < 0:
        inputClass = eval("onto." + inputCircuit.replace(' ', '_'))
    else:
        prefix = inputCircuit[:pos].strip().replace(' ', '_')
        inputCircuit = inputCircuit[pos + 1:].strip().replace(' ', '_')
        ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
        with ex_onto:
            inputClass = types.new_class(inputCircuit, (bifdns.Circuit,))
    cls.is_a.append(bifdns.inputCircuit.some(inputClass))
    pos = outputCircuit.find(':')
    if pos < 0:
        outputClass = eval("onto." + outputCircuit.replace(' ', '_'))
    else:
        prefix = outputCircuit[:pos].strip().replace(' ', '_')
        outputCircuit = outputCircuit[pos + 1:].strip().replace(' ', '_')
        ex_onto = get_ontology("https://wba-initiative.org/wbra/" + prefix + "/")
        with ex_onto:
            outputClass = types.new_class(outputCircuit, (bifdns.Circuit,))
    cls.is_a.append(bifdns.outputCircuit.some(outputClass))


if __name__ == '__main__':
    if len(sys.argv) <= 3:
        print("USE: python bif_exel2owl.py infile bifd_url outfile")
        exit()

    infilePath = sys.argv[1]
    outfilePath = sys.argv[3]
    lastSlash = infilePath.rfind('/')
    if lastSlash >= 0:
        infileName = infilePath[lastSlash + 1:]
    else:
        infileName = infilePath

    wb = openpyxl.load_workbook(sys.argv[1])
    bifd = get_ontology(sys.argv[2]).load()

    bifdns = bifd.get_namespace('https://wba-initiative.org/bifd/')
    # Defining annotation properties
    with bifdns:
        class functionality(comment):
            pass


        class reference(comment):
            pass


        class implementation(comment):
            pass

    # Defining an ontoloby
    project = wb['Project']
    pname = project.cell(row=2, column=1).value
    if not pname:
        print("Error: no project name")
        exit()
    description = project.cell(row=2, column=3).value
    onto = get_ontology("https://wba-initiative.org/wbra/" + pname + "/")
    onto.metadata.comment.append(description)

    createCircuits(wb['Circuit'])
    createConnections(wb['Connection'])
    # onto_path.append(outfilePath)
    onto.save(file=outfilePath)
